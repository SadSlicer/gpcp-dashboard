"""Data layer for the GPCP portfolio dashboard.

V5 — shares are no longer static: they are DERIVED from the transactions log
(BUY adds shares, SELL removes them), so holdings vary over time. A unitized
(time-weighted) VL isolates market performance from deposits/withdrawals. The
portfolio inception date is a user setting that drives every performance metric.

The Excel workbook (GPCP.xlsm) remains a **read-only seed**: it is imported once
into the sqlite sidecar (`portfolio.db`); all live state lives in sqlite
afterwards. (openpyxl can't recompute the workbook's formulas, so writing back
would corrupt it — hence sqlite.)
"""

from __future__ import annotations

import datetime as dt
import sqlite3
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
import pandas as pd

ROOT = Path(__file__).parent
WORKBOOK_PATH = ROOT / "GPCP.xlsm"
LEGACY_DB_PATH = ROOT / "portfolio.db"           # pre-V10 single-portfolio file
PORTFOLIOS_DIR = ROOT / "portfolios"             # V10: one sqlite per portfolio
REGISTRY_PATH = PORTFOLIOS_DIR / "_registry.json"

# --- Seed constants: the original 7 ETFs (used only to seed the DB once) -----
_SEED_ASSETS = [
    "S&P 500", "Emerging ESG", "Stoxx 600", "Russel 2000",
    "NASDAQ", "IBEX", "TOPIX",
]
_SEED_ISIN = {
    "S&P 500": "FR0011871128", "Emerging ESG": "FR0013412020",
    "Stoxx 600": "FR0011550193", "Russel 2000": "LU1681038672",
    "NASDAQ": "FR0011871110", "IBEX": "FR0010655746", "TOPIX": "FR0013411980",
}
_SEED_TICKER_BY_ISIN = {
    "FR0011871128": "PSP5.PA", "FR0013412020": "PAEEM.PA",
    "FR0011550193": "ETZ.PA", "LU1681038672": "RS2K.PA",
    "FR0011871110": "PUST.PA", "FR0010655746": "CS1.PA",
    "FR0013411980": "PTPXE.PA",
}

# --- Dynamic registry: kept in sync with the holdings table -----------------
# These module globals are mutated IN PLACE by _sync_registry() so that other
# modules which did `from data import ASSETS` see updates after a new ETF is
# added. Always read them after a data.* call (which runs ensure_seeded()).
ASSETS: list[str] = list(_SEED_ASSETS)
ISIN_BY_ASSET: dict[str, str] = dict(_SEED_ISIN)
YF_TICKER_BY_ISIN: dict[str, str] = dict(_SEED_TICKER_BY_ISIN)
# V10: direct asset→ticker map so assets without ISIN still resolve.
TICKER_BY_ASSET: dict[str, str] = {a: _SEED_TICKER_BY_ISIN[i] for a, i in _SEED_ISIN.items()}

# Workbook layout — used only when seeding sqlite from the .xlsm
MAJ_ISIN_ROW = 2
MAJ_FUND_ROW = 3
MAJ_FEES_ROW = 4
MAJ_SHARES_ROW = 8
MAJ_PRICE_HISTORY_FIRST_ROW = 27
PORTFOLIO_HISTORY_FIRST_ROW = 8


@dataclass
class PortfolioStatic:
    isins: dict[str, str] = field(default_factory=dict)
    funds: dict[str, str] = field(default_factory=dict)
    fees: dict[str, float] = field(default_factory=dict)
    shares: dict[str, float] = field(default_factory=dict)   # current, derived
    tickers: dict[str, str] = field(default_factory=dict)
    currencies: dict[str, str] = field(default_factory=dict)  # V11: native ccy per asset


# ---------------------------------------------------------------------------
# sqlite schema + bootstrap
# ---------------------------------------------------------------------------

SCHEMA = """
CREATE TABLE IF NOT EXISTS holdings (
    asset    TEXT PRIMARY KEY,
    isin     TEXT NOT NULL,
    fund     TEXT NOT NULL,
    fees     REAL NOT NULL,
    shares   INTEGER NOT NULL DEFAULT 0,   -- legacy/unused (shares derived from transactions)
    ticker   TEXT,
    added_at TEXT,
    currency TEXT NOT NULL DEFAULT 'EUR'   -- V11: native quote currency of this asset
);

-- V11: cached historical FX rates. `pair` is "XXXYYY" (e.g. USDEUR) meaning
-- 1 unit of XXX is worth `rate` units of YYY.
CREATE TABLE IF NOT EXISTS fx_rates (
    trade_date TEXT NOT NULL,
    pair       TEXT NOT NULL,
    rate       REAL NOT NULL,
    PRIMARY KEY (trade_date, pair)
);

CREATE TABLE IF NOT EXISTS prices (
    trade_date TEXT NOT NULL,
    asset      TEXT NOT NULL,
    price      REAL NOT NULL,
    PRIMARY KEY (trade_date, asset)
);

-- Each row is a BUY or SELL. shares > 0 always; txn_type carries direction.
-- amount_eur = price * shares (the cash that changed hands at execution).
CREATE TABLE IF NOT EXISTS transactions (
    id         INTEGER PRIMARY KEY AUTOINCREMENT,
    trade_date TEXT NOT NULL,
    asset      TEXT NOT NULL,
    isin       TEXT,
    txn_type   TEXT NOT NULL DEFAULT 'BUY',
    price      REAL,
    shares     REAL,
    amount_eur REAL NOT NULL
);

CREATE TABLE IF NOT EXISTS meta (
    key   TEXT PRIMARY KEY,
    value TEXT
);
"""


# ---------------------------------------------------------------------------
# V10 — multi-portfolio registry (one sqlite file per portfolio)
# ---------------------------------------------------------------------------

import json as _json
import re as _re


def _slugify(name: str) -> str:
    s = _re.sub(r"[^A-Za-z0-9]+", "_", name.strip()).strip("_")
    return s or "Portfolio"


def _default_registry() -> dict:
    return {"current_id": "GPCP", "portfolios": [
        {"id": "GPCP", "name": "GPCP", "created_at": "",
         "seed_from_workbook": True, "currency": "EUR"},
    ]}


def current_portfolio_currency() -> str:
    """Base currency of the active portfolio (default EUR for legacy)."""
    pf = current_portfolio()
    return (pf.get("currency") or "EUR").upper()


# In-process caches — V10 was hitting fd exhaustion because nav_series and
# compute_vl_series call shares_held_as_of() (which calls _all_transactions →
# _connect → _load_registry → fopen) once PER DATE in a loop. Caching the
# registry and the transaction list eliminates that per-iteration file pressure.
_REG_CACHE: dict | None = None
_REG_CACHE_MTIME: float = 0.0
_TX_CACHE: dict[str, list[dict]] = {}   # keyed by db path string


def _invalidate_caches() -> None:
    """Drop in-memory caches after any write (switch, create, delete, add txn)."""
    global _REG_CACHE, _REG_CACHE_MTIME
    _REG_CACHE = None
    _REG_CACHE_MTIME = 0.0
    _TX_CACHE.clear()


def _load_registry() -> dict:
    global _REG_CACHE, _REG_CACHE_MTIME
    PORTFOLIOS_DIR.mkdir(exist_ok=True)
    if not REGISTRY_PATH.exists():
        # Migrate from legacy single-portfolio layout:
        # if portfolio.db existed at the repo root, move it under portfolios/.
        reg = _default_registry()
        target = PORTFOLIOS_DIR / "portfolio_GPCP.db"
        if LEGACY_DB_PATH.exists() and not target.exists():
            LEGACY_DB_PATH.rename(target)
        reg["portfolios"][0]["created_at"] = dt.datetime.now().isoformat(timespec="seconds")
        REGISTRY_PATH.write_text(_json.dumps(reg, indent=2, ensure_ascii=False))
        _REG_CACHE, _REG_CACHE_MTIME = reg, REGISTRY_PATH.stat().st_mtime
        return reg
    try:
        mtime = REGISTRY_PATH.stat().st_mtime
    except OSError:
        mtime = 0.0
    if _REG_CACHE is not None and mtime == _REG_CACHE_MTIME:
        return _REG_CACHE
    reg = _json.loads(REGISTRY_PATH.read_text() or "{}") or _default_registry()
    # V11 backfill: legacy registry entries had no currency → default EUR
    dirty = False
    for p in reg.get("portfolios", []):
        if not p.get("currency"):
            p["currency"] = "EUR"
            dirty = True
    if dirty:
        REGISTRY_PATH.write_text(_json.dumps(reg, indent=2, ensure_ascii=False))
    _REG_CACHE, _REG_CACHE_MTIME = reg, REGISTRY_PATH.stat().st_mtime
    return reg


def _save_registry(reg: dict) -> None:
    global _REG_CACHE, _REG_CACHE_MTIME
    PORTFOLIOS_DIR.mkdir(exist_ok=True)
    REGISTRY_PATH.write_text(_json.dumps(reg, indent=2, ensure_ascii=False))
    _REG_CACHE = reg
    try:
        _REG_CACHE_MTIME = REGISTRY_PATH.stat().st_mtime
    except OSError:
        _REG_CACHE_MTIME = 0.0


def _current_db_path() -> Path:
    reg = _load_registry()
    cur = reg.get("current_id") or "GPCP"
    return PORTFOLIOS_DIR / f"portfolio_{cur}.db"


# ---- Public portfolio API (used by the Settings UI) ----

def list_portfolios() -> list[dict]:
    return _load_registry().get("portfolios", [])


def current_portfolio() -> dict:
    reg = _load_registry()
    cur = reg.get("current_id")
    for p in reg.get("portfolios", []):
        if p["id"] == cur:
            return p
    return reg["portfolios"][0]


def switch_portfolio(pid: str) -> None:
    reg = _load_registry()
    ids = {p["id"] for p in reg["portfolios"]}
    if pid not in ids:
        raise ValueError(f"Portefeuille inconnu : {pid}")
    reg["current_id"] = pid
    _save_registry(reg)
    # Reset in-memory registry so the next read pulls the new DB's holdings
    ASSETS[:] = []
    ISIN_BY_ASSET.clear()
    YF_TICKER_BY_ISIN.clear()
    _invalidate_caches()


def create_portfolio(name: str, currency: str = "EUR") -> dict:
    name = (name or "").strip()
    if not name:
        raise ValueError("Le nom du portefeuille est requis.")
    currency = (currency or "EUR").strip().upper() or "EUR"
    reg = _load_registry()
    pid = _slugify(name)
    existing = {p["id"] for p in reg["portfolios"]}
    base, n = pid, 2
    while pid in existing:
        pid = f"{base}_{n}"; n += 1
    entry = {"id": pid, "name": name, "seed_from_workbook": False,
             "currency": currency,
             "created_at": dt.datetime.now().isoformat(timespec="seconds")}
    reg["portfolios"].append(entry)
    _save_registry(reg)
    # Pre-create the sqlite file with the schema and mark it as seeded so
    # ensure_seeded() does NOT pull the workbook (this portfolio starts empty).
    target = PORTFOLIOS_DIR / f"portfolio_{pid}.db"
    con = sqlite3.connect(target)
    con.row_factory = sqlite3.Row
    try:
        con.executescript(SCHEMA)
        _migrate(con)
        con.execute("INSERT OR REPLACE INTO meta(key,value) VALUES('seeded','1')")
        con.execute("INSERT OR REPLACE INTO meta(key,value) VALUES('seeded_at',?)",
                    (dt.datetime.now().isoformat(timespec="seconds"),))
        con.commit()
    finally:
        con.close()
    _invalidate_caches()
    return entry


def rename_portfolio(pid: str, new_name: str) -> None:
    new_name = (new_name or "").strip()
    if not new_name:
        raise ValueError("Nom invalide.")
    reg = _load_registry()
    for p in reg["portfolios"]:
        if p["id"] == pid:
            p["name"] = new_name
            _save_registry(reg)
            return
    raise ValueError(f"Portefeuille inconnu : {pid}")


def delete_portfolio(pid: str) -> None:
    reg = _load_registry()
    if len(reg["portfolios"]) <= 1:
        raise ValueError("Impossible : il doit rester au moins un portefeuille.")
    keep = [p for p in reg["portfolios"] if p["id"] != pid]
    if len(keep) == len(reg["portfolios"]):
        raise ValueError(f"Portefeuille inconnu : {pid}")
    reg["portfolios"] = keep
    if reg.get("current_id") == pid:
        reg["current_id"] = keep[0]["id"]
        ASSETS[:] = []; ISIN_BY_ASSET.clear(); YF_TICKER_BY_ISIN.clear()
    _save_registry(reg)
    (PORTFOLIOS_DIR / f"portfolio_{pid}.db").unlink(missing_ok=True)
    _invalidate_caches()


def _connect() -> sqlite3.Connection:
    con = sqlite3.connect(_current_db_path())
    con.row_factory = sqlite3.Row
    return con


def _columns(con: sqlite3.Connection, table: str) -> set[str]:
    return {r["name"] for r in con.execute(f"PRAGMA table_info({table})")}


def _ensure_db() -> None:
    with _connect() as con:
        con.executescript(SCHEMA)
        _migrate(con)
        con.commit()


def _migrate(con: sqlite3.Connection) -> None:
    """Bring an old (V4) DB up to the V5 schema, in place, idempotently."""
    # holdings: add ticker / added_at / currency if missing
    hcols = _columns(con, "holdings")
    if "ticker" not in hcols:
        con.execute("ALTER TABLE holdings ADD COLUMN ticker TEXT")
    if "added_at" not in hcols:
        con.execute("ALTER TABLE holdings ADD COLUMN added_at TEXT")
    if "currency" not in hcols:
        # V11: default every legacy asset to EUR (true for the original 7 ETFs
        # on Euronext Paris and for any V5–V10 user-added EUR ETF).
        con.execute("ALTER TABLE holdings ADD COLUMN currency TEXT NOT NULL DEFAULT 'EUR'")
    # V11: fx_rates table (idempotent CREATE so old DBs get it too)
    con.execute("""
        CREATE TABLE IF NOT EXISTS fx_rates (
            trade_date TEXT NOT NULL,
            pair       TEXT NOT NULL,
            rate       REAL NOT NULL,
            PRIMARY KEY (trade_date, pair)
        )
    """)
    # backfill tickers from seed map
    for r in con.execute("SELECT asset, isin, ticker FROM holdings").fetchall():
        if not r["ticker"]:
            tk = _SEED_TICKER_BY_ISIN.get(r["isin"])
            if tk:
                con.execute("UPDATE holdings SET ticker=? WHERE asset=?", (tk, r["asset"]))

    # transactions: V4 had (trade_date, asset, isin, amount_eur) with no id/type.
    tcols = _columns(con, "transactions")
    needs_rebuild = "id" not in tcols
    if needs_rebuild:
        # Recreate with the new schema, preserving old rows.
        old = con.execute(
            "SELECT trade_date, asset, isin, amount_eur FROM transactions"
        ).fetchall()
        con.execute("ALTER TABLE transactions RENAME TO _transactions_old")
        con.execute("""
            CREATE TABLE transactions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                trade_date TEXT NOT NULL, asset TEXT NOT NULL, isin TEXT,
                txn_type TEXT NOT NULL DEFAULT 'BUY', price REAL, shares REAL,
                amount_eur REAL NOT NULL
            )
        """)
        # seed shares (legacy holdings.shares) used to backfill price = amount/shares
        seed_shares = {r["asset"]: r["shares"]
                       for r in con.execute("SELECT asset, shares FROM holdings")}
        for r in old:
            sh = seed_shares.get(r["asset"]) or 0
            price = (r["amount_eur"] / sh) if sh else None
            con.execute(
                "INSERT INTO transactions(trade_date,asset,isin,txn_type,price,shares,amount_eur)"
                " VALUES (?,?,?,?,?,?,?)",
                (r["trade_date"], r["asset"], r["isin"], "BUY", price, float(sh), r["amount_eur"]),
            )
        con.execute("DROP TABLE _transactions_old")
    else:
        # Already new schema; just ensure the optional columns exist.
        for col, decl in [("txn_type", "TEXT NOT NULL DEFAULT 'BUY'"),
                          ("price", "REAL"), ("shares", "REAL")]:
            if col not in tcols:
                con.execute(f"ALTER TABLE transactions ADD COLUMN {col} {decl}")

    # V12: per-transaction currency. Default EUR; legacy rows are backfilled
    # from the asset's holdings.currency (which itself defaults to EUR).
    tcols = _columns(con, "transactions")
    if "currency" not in tcols:
        con.execute(
            "ALTER TABLE transactions ADD COLUMN currency TEXT NOT NULL DEFAULT 'EUR'"
        )
        con.execute("""
            UPDATE transactions
               SET currency = COALESCE(
                       (SELECT h.currency FROM holdings h
                         WHERE h.asset = transactions.asset),
                       'EUR'
                   )
        """)


def _sync_registry(con: sqlite3.Connection) -> None:
    """Refresh the module-global ASSETS / ISIN / TICKER maps from holdings.

    Order: the 7 seed assets first (in canonical order), then any new assets
    by added_at. Mutates the globals in place so imported references update.
    """
    rows = con.execute("SELECT asset, isin, ticker, added_at FROM holdings").fetchall()
    by_asset = {r["asset"]: r for r in rows}
    seed_present = [a for a in _SEED_ASSETS if a in by_asset]
    extras = [a for a in by_asset if a not in _SEED_ASSETS]
    extras.sort(key=lambda a: (by_asset[a]["added_at"] or "", a))
    ordered = seed_present + extras

    ASSETS[:] = ordered
    ISIN_BY_ASSET.clear()
    YF_TICKER_BY_ISIN.clear()
    TICKER_BY_ASSET.clear()
    for a in ordered:
        r = by_asset[a]
        ISIN_BY_ASSET[a] = r["isin"] or ""
        if r["ticker"]:
            TICKER_BY_ASSET[a] = r["ticker"]
        if r["isin"] and r["ticker"]:
            YF_TICKER_BY_ISIN[r["isin"]] = r["ticker"]


# ---------------------------------------------------------------------------
# One-time seed from the Excel workbook
# ---------------------------------------------------------------------------

def _read_workbook_into_seed() -> dict:
    wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    maj = wb["MAJ"]

    static = PortfolioStatic()
    for col_idx, asset in enumerate(_SEED_ASSETS, start=2):
        static.isins[asset] = maj.cell(row=MAJ_ISIN_ROW, column=col_idx).value
        static.funds[asset] = maj.cell(row=MAJ_FUND_ROW, column=col_idx).value
        static.fees[asset] = maj.cell(row=MAJ_FEES_ROW, column=col_idx).value or 0.0
        shares = maj.cell(row=MAJ_SHARES_ROW, column=col_idx).value
        static.shares[asset] = int(shares) if shares is not None else 0

    price_rows: list[tuple[dt.date, str, float]] = []
    for row in maj.iter_rows(min_row=MAJ_PRICE_HISTORY_FIRST_ROW, values_only=True):
        d = row[0]
        if not isinstance(d, dt.datetime):
            break
        vals = row[1:1 + len(_SEED_ASSETS)]
        if all(v is None or v == 0 for v in vals):
            continue
        for asset, v in zip(_SEED_ASSETS, vals):
            if v is None:
                continue
            try:
                price_rows.append((d.date(), asset, float(v)))
            except (TypeError, ValueError):
                continue

    for r in (9, 10):
        d = maj.cell(row=r, column=1).value
        if isinstance(d, dt.datetime):
            for col_idx, asset in enumerate(_SEED_ASSETS, start=2):
                v = maj.cell(row=r, column=col_idx).value
                if isinstance(v, (int, float)) and v != 0:
                    price_rows.append((d.date(), asset, float(v)))

    pf = wb["Portfolio"]
    for row in pf.iter_rows(min_row=PORTFOLIO_HISTORY_FIRST_ROW, values_only=True):
        d = row[0]
        if not isinstance(d, dt.datetime):
            break
        vals = row[1:1 + len(_SEED_ASSETS)]
        if all(v is None or v == 0 for v in vals):
            continue
        for asset, v in zip(_SEED_ASSETS, vals):
            sh = static.shares.get(asset, 0)
            if v is None or sh == 0:
                continue
            try:
                price_rows.append((d.date(), asset, float(v) / sh))
            except (TypeError, ValueError):
                continue

    tx_rows: list[tuple[dt.date, str, str | None, float]] = []
    for row in wb["Transactions"].iter_rows(min_row=4, values_only=True):
        if row[0] is None:
            continue
        d, asset, isin, amount = row[:4]
        if not isinstance(d, dt.datetime):
            continue
        if isinstance(asset, str):
            asset = asset.strip()
        tx_rows.append((d.date(), asset, isin, float(amount) if amount is not None else 0.0))

    return {"static": static, "prices": price_rows, "transactions": tx_rows}


def _is_seeded(con: sqlite3.Connection) -> bool:
    row = con.execute("SELECT value FROM meta WHERE key='seeded'").fetchone()
    return bool(row and row["value"] == "1")


def ensure_seeded() -> None:
    """Create + migrate the DB, import from the workbook if empty, sync registry.

    The workbook seed only runs for portfolios flagged `seed_from_workbook` in
    the registry (only GPCP). Other portfolios start empty (schema only).
    """
    _ensure_db()
    cur_pf = current_portfolio()
    with _connect() as con:
        if not _is_seeded(con) and cur_pf.get("seed_from_workbook"):
            seed = _read_workbook_into_seed()
            static: PortfolioStatic = seed["static"]
            now = dt.datetime.now().isoformat(timespec="seconds")
            for a in _SEED_ASSETS:
                con.execute(
                    "INSERT OR REPLACE INTO holdings(asset,isin,fund,fees,shares,ticker,added_at)"
                    " VALUES (?,?,?,?,?,?,?)",
                    (a, static.isins[a], static.funds[a], static.fees[a],
                     int(static.shares[a]), _SEED_TICKER_BY_ISIN.get(static.isins[a]), now),
                )
            dedup: dict[tuple[dt.date, str], float] = {}
            for d, a, p in seed["prices"]:
                dedup[(d, a)] = p
            con.executemany(
                "INSERT OR REPLACE INTO prices(trade_date,asset,price) VALUES (?,?,?)",
                [(d.isoformat(), a, p) for (d, a), p in dedup.items()],
            )
            # Seed transactions as BUYs, backfilling price = amount / seed shares.
            for d, a, isin, amt in seed["transactions"]:
                sh = static.shares.get(a, 0)
                price = (amt / sh) if sh else None
                con.execute(
                    "INSERT INTO transactions(trade_date,asset,isin,txn_type,price,shares,amount_eur)"
                    " VALUES (?,?,?,?,?,?,?)",
                    (d.isoformat(), a, isin, "BUY", price, float(sh), amt),
                )
            con.execute("INSERT OR REPLACE INTO meta(key,value) VALUES('seeded','1')")
            con.execute("INSERT OR REPLACE INTO meta(key,value) VALUES(?,?)",
                        ("seeded_at", now))
            con.commit()
        elif not _is_seeded(con):
            # Non-workbook portfolio (empty by design) — just mark seeded.
            con.execute("INSERT OR REPLACE INTO meta(key,value) VALUES('seeded','1')")
            con.execute(
                "INSERT OR REPLACE INTO meta(key,value) VALUES(?,?)",
                ("seeded_at", dt.datetime.now().isoformat(timespec="seconds")),
            )
            con.commit()
        _sync_registry(con)


def reset_db_and_reseed() -> None:
    """Reset the CURRENT portfolio's DB. For GPCP this re-seeds from the
    workbook; for any other portfolio this just recreates an empty DB."""
    db = _current_db_path()
    if db.exists():
        db.unlink()
    ensure_seeded()


# ---------------------------------------------------------------------------
# Settings: portfolio inception date
# ---------------------------------------------------------------------------

def get_inception_date() -> dt.date:
    """The portfolio start date. Default = earliest transaction date (else
    earliest price date, else today)."""
    ensure_seeded()
    with _connect() as con:
        row = con.execute("SELECT value FROM meta WHERE key='inception_date'").fetchone()
        if row and row["value"]:
            try:
                return dt.date.fromisoformat(row["value"])
            except ValueError:
                pass
        r = con.execute("SELECT MIN(trade_date) AS d FROM transactions").fetchone()
        if r and r["d"]:
            return dt.date.fromisoformat(r["d"][:10])
        r = con.execute("SELECT MIN(trade_date) AS d FROM prices").fetchone()
        if r and r["d"]:
            return dt.date.fromisoformat(r["d"][:10])
    return dt.date.today()


def set_inception_date(d: dt.date) -> None:
    ensure_seeded()
    with _connect() as con:
        con.execute("INSERT OR REPLACE INTO meta(key,value) VALUES('inception_date',?)",
                    (d.isoformat(),))
        con.commit()


def default_inception_date() -> dt.date:
    """Earliest transaction date, ignoring any manual override."""
    ensure_seeded()
    with _connect() as con:
        r = con.execute("SELECT MIN(trade_date) AS d FROM transactions").fetchone()
    if r and r["d"]:
        return dt.date.fromisoformat(r["d"][:10])
    return dt.date.today()


# ---------------------------------------------------------------------------
# Shares (derived from the transactions log)
# ---------------------------------------------------------------------------

def _all_transactions() -> list[dict]:
    """Cached: returns the list of transactions for the CURRENT portfolio's DB.

    The cache is keyed by the DB file path; it's invalidated on any portfolio
    switch / write / add_transaction (see `_invalidate_caches`)."""
    key = str(_current_db_path())
    cached = _TX_CACHE.get(key)
    if cached is not None:
        return cached
    ensure_seeded()
    with _connect() as con:
        rows = con.execute(
            "SELECT trade_date, asset, isin, txn_type, price, shares, amount_eur,"
            " currency FROM transactions ORDER BY trade_date, id"
        ).fetchall()
    out = []
    for r in rows:
        out.append({
            "date": dt.date.fromisoformat(r["trade_date"][:10]),
            "asset": r["asset"], "isin": r["isin"],
            "type": (r["txn_type"] or "BUY").upper(),
            "price": r["price"], "shares": float(r["shares"] or 0),
            "amount": float(r["amount_eur"] or 0),
            "currency": (r["currency"] or "EUR").upper(),
        })
    _TX_CACHE[key] = out
    return out


def shares_held_as_of(when: dt.date | None = None) -> dict[str, float]:
    """Net shares per asset from all transactions with date <= `when`."""
    when = when or dt.date.today()
    held: dict[str, float] = defaultdict(float)
    for t in _all_transactions():
        if t["date"] <= when and t["type"] in ("BUY", "SELL"):
            held[t["asset"]] += t["shares"] if t["type"] == "BUY" else -t["shares"]
    return {a: s for a, s in held.items()}


def current_shares() -> dict[str, float]:
    return shares_held_as_of(dt.date.today())


# ---------------------------------------------------------------------------
# Cash account (PEA model: sells become cash; buys draw cash, shortfall = deposit)
# ---------------------------------------------------------------------------

def _cash_walk(upto: dt.date | None = None) -> dict:
    """Walk transactions chronologically maintaining a cash balance.

    PEA logic:
      • SELL    → proceeds added to cash (internal, no external flow)
      • BUY     → funded from cash first; any shortfall is a fresh EXTERNAL
                  deposit (new money in)
      • DEPOSIT → explicit external cash in  (cash +, external deposit +)
      • WITHDRAW→ explicit external cash out (cash −, external deposit −)

    `deposit_by_date` holds the NET external flow per date (deposits positive,
    withdrawals negative) — this is what drives unit creation/destruction in the
    unitized VL. Returns: cash, deposits (cumulative net), cash_by_date,
    deposit_by_date.
    """
    # V12: every transaction carries its OWN currency (the currency that was
    # actually debited/credited from the broker's cash account). Convert that
    # native amount to portfolio currency via FX at the transaction date.
    # Cash itself always lives in portfolio currency. DEPOSIT/WITHDRAW are
    # already stored in pf ccy (the form's "Montant" field is in pf ccy).
    pf_ccy = current_portfolio_currency()

    cash = 0.0
    deposits = 0.0
    cash_by_date: dict[dt.date, float] = {}
    deposit_by_date: dict[dt.date, float] = defaultdict(float)
    for t in _all_transactions():
        if upto is not None and t["date"] > upto:
            break
        amt = t["amount"]
        typ = t["type"]
        # Convert BUY/SELL native amounts (in t["currency"]) to portfolio ccy.
        if typ in ("BUY", "SELL"):
            tx_ccy = (t.get("currency") or "EUR").upper()
            if tx_ccy != pf_ccy:
                amt = amt * fx_rate(tx_ccy, pf_ccy, t["date"])
        if typ == "BUY":
            if cash >= amt:
                cash -= amt
            else:
                deposit = amt - cash
                deposits += deposit
                deposit_by_date[t["date"]] += deposit
                cash = 0.0
        elif typ == "SELL":
            cash += amt
        elif typ == "DEPOSIT":
            cash += amt
            deposits += amt
            deposit_by_date[t["date"]] += amt
        elif typ == "WITHDRAW":
            cash -= amt
            deposits -= amt
            deposit_by_date[t["date"]] -= amt
        cash_by_date[t["date"]] = cash
    return {"cash": cash, "deposits": deposits,
            "cash_by_date": cash_by_date, "deposit_by_date": dict(deposit_by_date)}


def cash_balance_as_of(when: dt.date | None = None) -> float:
    return _cash_walk(when or dt.date.today())["cash"]


def external_deposits_as_of(when: dt.date | None = None) -> float:
    return _cash_walk(when or dt.date.today())["deposits"]


def nav_series(price_history: pd.DataFrame | None = None,
               inception: dt.date | None = None) -> pd.DataFrame:
    """Per-date portfolio NAV = ETF market value (time-varying shares) + cash.

    Returns columns: date, etf, cash, nav. Used by the VL engine, the Overview
    NAV chart and the Pro tab so NAV is consistent everywhere (incl. cash).
    """
    if price_history is None:
        price_history = load_price_history(full=(inception is None))
    inception = inception or get_inception_date()
    if price_history.empty:
        return pd.DataFrame(columns=["date", "etf", "cash", "nav"])
    ph = price_history[price_history["date"].dt.date >= inception].sort_values("date").reset_index(drop=True)
    if ph.empty:
        return pd.DataFrame(columns=["date", "etf", "cash", "nav"])
    assets = [a for a in ASSETS if a in ph.columns]
    walk = _cash_walk()
    cash_by_date = walk["cash_by_date"]

    # V11: convert each asset's native price to the portfolio's base currency.
    pf_ccy = current_portfolio_currency()
    static = load_static()
    asset_ccy = {a: (static.currencies.get(a) or "EUR").upper() for a in assets}
    # Pre-warm FX cache for each non-pf currency over the chart window
    if not ph.empty:
        d0 = ph.at[0, "date"].date()
        d1 = ph.at[len(ph) - 1, "date"].date()
        for ccy in set(asset_ccy.values()):
            if ccy != pf_ccy:
                try:
                    prefetch_fx_window(ccy, pf_ccy, d0, d1)
                except Exception:
                    pass

    def cash_at(d: dt.date) -> float:
        applicable = [cd for cd in cash_by_date if cd <= d]
        return cash_by_date[max(applicable)] if applicable else 0.0

    rows = []
    for i in range(len(ph)):
        d = ph.at[i, "date"].date()
        held = shares_held_as_of(d)
        etf = 0.0
        for a in assets:
            v = ph.at[i, a]
            if not pd.notna(v):
                continue
            rate = 1.0 if asset_ccy[a] == pf_ccy else fx_rate(asset_ccy[a], pf_ccy, d)
            etf += float(v) * held.get(a, 0) * rate
        cash = cash_at(d)
        rows.append({"date": ph.at[i, "date"], "etf": etf, "cash": cash, "nav": etf + cash})
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Read APIs
# ---------------------------------------------------------------------------

def load_static() -> PortfolioStatic:
    """Per-asset metadata + CURRENT (today) shares derived from transactions."""
    ensure_seeded()
    with _connect() as con:
        rows = con.execute("SELECT * FROM holdings").fetchall()
    by_asset = {r["asset"]: r for r in rows}
    shares_now = current_shares()
    static = PortfolioStatic()
    for asset in ASSETS:
        r = by_asset.get(asset)
        if r is None:
            continue
        static.isins[asset] = r["isin"]
        static.funds[asset] = r["fund"]
        static.fees[asset] = float(r["fees"])
        static.tickers[asset] = r["ticker"]
        static.shares[asset] = float(shares_now.get(asset, 0))
        # V11: currency may be missing on rows from a freshly-migrated DB
        try:
            static.currencies[asset] = (r["currency"] or "EUR").upper()
        except (IndexError, KeyError):
            static.currencies[asset] = "EUR"
    return static


def load_price_history(full: bool = False) -> pd.DataFrame:
    """Wide price history: one row per date, one column per asset.

    By default returns only dates >= the portfolio inception (so every UI /
    performance consumer respects the Settings inception date). Pass full=True
    to get the entire table (used by internal/price-feed operations).
    """
    ensure_seeded()
    with _connect() as con:
        df = pd.read_sql_query("SELECT trade_date, asset, price FROM prices", con)
    if df.empty:
        return pd.DataFrame(columns=["date", *ASSETS])
    wide = df.pivot(index="trade_date", columns="asset", values="price").reset_index()
    wide = wide.rename(columns={"trade_date": "date"})
    wide["date"] = pd.to_datetime(wide["date"])
    for a in ASSETS:
        if a not in wide.columns:
            wide[a] = pd.NA
    wide = wide[["date", *ASSETS]].sort_values("date").reset_index(drop=True)
    if not full:
        inception = get_inception_date()
        wide = wide[wide["date"].dt.date >= inception].reset_index(drop=True)
    return wide


def load_position_history(full: bool = False) -> pd.DataFrame:
    """Wide position-value history using TIME-VARYING shares held at each date."""
    prices = load_price_history(full=full)
    if prices.empty:
        return pd.DataFrame(columns=["date", *ASSETS])
    out = prices.copy()
    # Precompute shares per date (rows are sorted ascending)
    for idx, row in out.iterrows():
        d = row["date"].date()
        held = shares_held_as_of(d)
        for a in ASSETS:
            price = prices.at[idx, a]
            out.at[idx, a] = (float(price) * held.get(a, 0)) if pd.notna(price) else pd.NA
    return out


def price_history_in_portfolio_currency(price_history: pd.DataFrame) -> pd.DataFrame:
    """V12: return a copy of price_history with each asset column converted
    from its NATIVE currency to the current portfolio currency, using
    fx_rate at each row's date. EUR-on-EUR assets are no-ops.
    """
    if price_history is None or price_history.empty:
        return price_history.copy() if price_history is not None else pd.DataFrame()
    out = price_history.copy()
    pf_ccy = current_portfolio_currency()
    static = load_static()
    asset_ccy = {a: (static.currencies.get(a) or "EUR").upper() for a in ASSETS}
    # Pre-warm FX cache for each non-pf currency over the window
    d0 = out["date"].min().date()
    d1 = out["date"].max().date()
    for ccy in set(asset_ccy.values()):
        if ccy != pf_ccy:
            try:
                prefetch_fx_window(ccy, pf_ccy, d0, d1)
            except Exception:
                pass
    for a in ASSETS:
        if a not in out.columns:
            continue
        ccy = asset_ccy[a]
        if ccy == pf_ccy:
            continue
        rates = out["date"].apply(
            lambda ts: fx_rate(ccy, pf_ccy, ts.date() if hasattr(ts, "date") else ts)
        )
        out[a] = out[a].astype(float) * rates.astype(float)
    return out


def trading_day_rangebreaks(price_history: pd.DataFrame | None = None,
                            start: dt.date | None = None,
                            end: dt.date | None = None) -> list[dict]:
    """V12: Plotly `rangebreaks` config to hide non-trading days on time-series
    charts. Skips weekends (Sat/Sun) and any weekday absent from the price
    history (= market holidays). Pass the price_history slice that feeds the
    chart, or None to skip the holiday list and only break weekends.
    """
    breaks: list[dict] = [dict(bounds=["sat", "mon"])]
    if price_history is None or price_history.empty:
        return breaks
    trading = set(pd.to_datetime(price_history["date"]).dt.date)
    if start is None:
        start = min(trading)
    if end is None:
        end = max(trading)
    weekdays = pd.bdate_range(start, end).date
    holidays = [d.isoformat() for d in weekdays if d not in trading]
    if holidays:
        breaks.append(dict(values=holidays))
    return breaks


def load_transactions() -> pd.DataFrame:
    """Enriched transaction history for the UI.

    V12: `Total` is the NATIVE-currency amount (price × shares in the tx's
    currency) and `Currency` is that currency. UI converts to portfolio
    currency at display time using fx_rate at the trade date.
    V15: returns the `id` column so the UI can reference rows for deletion.
    """
    ensure_seeded()
    with _connect() as con:
        df = pd.read_sql_query(
            "SELECT id AS Id, trade_date AS Date, asset AS Asset, isin AS ISIN,"
            " txn_type AS Type, price AS Price, shares AS Shares,"
            " amount_eur AS Total, currency AS Currency FROM transactions",
            con,
        )
    if df.empty:
        return df
    df["Date"] = pd.to_datetime(df["Date"])
    df["Currency"] = df["Currency"].fillna("EUR").str.upper()
    return df.sort_values("Date", ascending=False).reset_index(drop=True)


def add_transactions_bulk(rows: list[dict]) -> dict:
    """VA2: bulk-insert a list of transaction dicts in a single
    transaction. Each row dict must have keys:
        - date: dt.date
        - asset: str (must be in ASSETS, except for cash movements
                       where asset='Cash')
        - type: 'BUY' | 'SELL' | 'DEPOSIT' | 'WITHDRAW'
        - shares: float (for BUY/SELL) — required
        - price: float (for BUY/SELL) — optional, auto-fetched if None
        - currency: str (for BUY/SELL) — optional, defaults to asset's ccy
        - amount: float (for DEPOSIT/WITHDRAW) — required
    Returns {"inserted": int, "skipped": int, "errors": [...]}.

    Skipped reasons: missing required field, unknown asset, price not
    found and no manual override, oversell, etc. Errors carry the row
    index + reason.
    """
    if not rows:
        return {"inserted": 0, "skipped": 0, "errors": []}
    ensure_seeded()
    inserted = 0
    errors: list[dict] = []
    pf_ccy = current_portfolio_currency()
    for i, r in enumerate(rows):
        try:
            t = (r.get("type") or "").strip().upper()
            d = r.get("date")
            if isinstance(d, str):
                d = dt.date.fromisoformat(d)
            if not isinstance(d, dt.date):
                raise ValueError("date invalide")
            if t in ("DEPOSIT", "WITHDRAW"):
                amount = float(r.get("amount") or 0)
                if amount <= 0:
                    raise ValueError("amount doit être > 0")
                add_cash_movement(d, t, amount)
            elif t in ("BUY", "SELL"):
                asset = (r.get("asset") or "").strip()
                if asset not in ASSETS:
                    raise ValueError(f"asset inconnu : {asset!r}")
                shares = float(r.get("shares") or 0)
                if shares <= 0:
                    raise ValueError("shares doit être > 0")
                price = r.get("price")
                if price is None or float(price) <= 0:
                    auto_p, _src = price_on_date(asset, d)
                    if auto_p is None:
                        raise ValueError("prix introuvable et non fourni")
                    price = auto_p
                price = float(price)
                ccy = (r.get("currency") or "").strip().upper()
                # Default currency = asset's registered currency
                if not ccy:
                    with _connect() as con:
                        row = con.execute(
                            "SELECT currency FROM holdings WHERE asset=?",
                            (asset,),
                        ).fetchone()
                    ccy = ((row["currency"] if row and row["currency"]
                            else pf_ccy) or pf_ccy).upper()
                add_transaction(d, asset, t, price=price, shares=shares,
                                currency=ccy)
            else:
                raise ValueError(f"type inconnu : {t!r}")
            inserted += 1
        except Exception as exc:
            errors.append({"row": i, "reason": str(exc), "data": r})
    _invalidate_caches()
    return {"inserted": inserted,
            "skipped": len(rows) - inserted,
            "errors": errors}


def delete_transaction(tx_id: int) -> dict:
    """V15: delete one transaction by id. Idempotent on a missing id.

    Doesn't reverse any cash movement explicitly — the cash account
    (`_cash_walk`) and shares (`shares_held_as_of`) are *derived* from
    the full transaction history on every read, so removing a row makes
    everything re-converge automatically. Caches are invalidated so the
    next snapshot/nav/VL is built from the post-delete state.

    Returns: {"deleted": int, "id": int, "row": dict | None}
    """
    ensure_seeded()
    tx_id = int(tx_id)
    with _connect() as con:
        row = con.execute(
            "SELECT id, trade_date, asset, txn_type, price, shares, amount_eur,"
            " currency FROM transactions WHERE id=?",
            (tx_id,),
        ).fetchone()
        if row is None:
            return {"deleted": 0, "id": tx_id, "row": None}
        cur = con.execute("DELETE FROM transactions WHERE id=?", (tx_id,))
        con.commit()
        deleted = cur.rowcount
    _invalidate_caches()
    return {
        "deleted": int(deleted),
        "id": tx_id,
        "row": {k: row[k] for k in row.keys()} if row else None,
    }


# ---------------------------------------------------------------------------
# Write APIs — transactions & assets
# ---------------------------------------------------------------------------

def register_asset(asset: str, isin: str, ticker: str,
                   fund: str = "—", fees: float = 0.0,
                   currency: str = "EUR") -> None:
    """Add a brand-new asset to the registry. Idempotent upsert.

    V11: `currency` is the native quote currency of this asset (e.g. USD for AAPL).
    The dashboard converts to the active portfolio's currency automatically.
    """
    ensure_seeded()
    currency = (currency or "EUR").strip().upper() or "EUR"
    now = dt.datetime.now().isoformat(timespec="seconds")
    with _connect() as con:
        con.execute(
            "INSERT OR IGNORE INTO holdings(asset,isin,fund,fees,shares,ticker,added_at,currency)"
            " VALUES (?,?,?,?,?,?,?,?)",
            (asset, isin, fund, fees, 0, ticker, now, currency),
        )
        con.execute(
            "UPDATE holdings SET isin=?, ticker=?, fund=?, fees=?, currency=? WHERE asset=?",
            (isin, ticker, fund, fees, currency, asset),
        )
        con.commit()
        _sync_registry(con)
    # Make sure the new ETF gets a composition entry so the monthly scraper
    # (and the immediate scrape from the UI) treat it like every other ETF.
    try:
        import compositions_scraper as _cs
        _cs.ensure_asset_entry(asset, isin, ticker, issuer=fund)
    except Exception:
        pass
    # Give the new ETF a stub in the compositions registry so the monthly
    # scraper treats it like every other ETF. Best-effort, never blocks.
    try:
        import compositions_scraper as _cs
        _cs.ensure_asset_entry(asset, isin, ticker, issuer=fund)
    except Exception:
        pass


def add_transaction(when: dt.date, asset: str, txn_type: str,
                    price: float, shares: float, isin: str | None = None,
                    currency: str | None = None) -> dict:
    """Record a BUY or SELL. Blocks oversell. Returns a summary.

    V12: `currency` is the currency in which `price` is denominated (= the
    currency that was debited/credited from your broker's cash). Defaults to
    the asset's registered currency.
    """
    ensure_seeded()
    txn_type = txn_type.upper()
    if txn_type not in ("BUY", "SELL"):
        raise ValueError("txn_type must be BUY or SELL")
    if shares <= 0 or price < 0:
        raise ValueError("shares must be > 0 and price >= 0")

    if txn_type == "SELL":
        held = shares_held_as_of(when).get(asset, 0)
        if shares > held + 1e-9:
            raise ValueError(
                f"Vente refusée : tu détiens {held:g} parts de {asset} au {when.isoformat()}, "
                f"impossible d'en vendre {shares:g}."
            )

    if isin is None:
        isin = ISIN_BY_ASSET.get(asset)
    if not currency:
        # V12: default to the asset's registered native currency
        try:
            static = load_static()
            currency = (static.currencies.get(asset) or "EUR").upper()
        except Exception:
            currency = "EUR"
    currency = currency.upper()
    amount = round(price * shares, 2)
    with _connect() as con:
        con.execute(
            "INSERT INTO transactions(trade_date,asset,isin,txn_type,price,shares,amount_eur,currency)"
            " VALUES (?,?,?,?,?,?,?,?)",
            (when.isoformat(), asset, isin, txn_type, price, float(shares), amount, currency),
        )
        con.commit()
    _invalidate_caches()

    # Auto-backfill: if the transaction predates the asset's earliest saved
    # price, fetch the missing daily closes from Yahoo so the position can be
    # valued back to this date. Best-effort; failures are silent (the daily
    # cron will keep trying anyway).
    backfilled = 0
    try:
        backfilled = _backfill_asset_since(asset, when)
    except Exception:
        pass

    return {"date": when.isoformat(), "asset": asset, "type": txn_type,
            "price": price, "shares": shares, "amount": amount,
            "new_holding": shares_held_as_of(dt.date.today()).get(asset, 0),
            "backfilled_days": backfilled}


def backfill_all_assets_to_inception() -> dict:
    """For every registered asset, fetch and store any missing Yahoo daily
    closes between the portfolio inception date and today. Used by the
    Settings → "Rétro-importer les prix" button to fix portfolios whose
    transactions predate their saved prices.

    Returns: {asset: rows_inserted} (assets with 0 are omitted).
    """
    ensure_seeded()
    inception = get_inception_date()
    out: dict[str, int] = {}
    for asset in list(ASSETS):
        try:
            n = _backfill_asset_since(asset, inception)
        except Exception:
            n = 0
        if n:
            out[asset] = n
    return out


def refetch_all_prices_from_inception() -> dict:
    """V12: wipe the prices table and refetch every asset's full history from
    inception, using `auto_adjust=True` (adjusted close). Use this once after
    upgrading to V12 to drop any legacy raw-close prices that pre-dated the
    auto_adjust=True switch, so the whole series is internally consistent
    (split- and dividend-adjusted total-return).

    Returns: {asset: rows_inserted}.
    """
    ensure_seeded()
    with _connect() as con:
        con.execute("DELETE FROM prices")
        con.commit()
    _invalidate_caches()
    inception = get_inception_date()
    out: dict[str, int] = {}
    for asset in list(ASSETS):
        try:
            n = _backfill_asset_since(asset, inception)
        except Exception:
            n = 0
        out[asset] = n
    return out


def _backfill_asset_since(asset: str, since: dt.date) -> int:
    """Fetch and store this asset's Yahoo daily closes from `since` until today
    for any dates not already present. Returns the number of rows inserted."""
    ticker = TICKER_BY_ASSET.get(asset)
    if not ticker:
        return 0
    # Skip if we already have coverage going back at least to `since`.
    with _connect() as con:
        r = con.execute("SELECT MIN(trade_date) AS d FROM prices WHERE asset=?",
                        (asset,)).fetchone()
    earliest = dt.date.fromisoformat(r["d"][:10]) if r and r["d"] else None
    if earliest and earliest <= since:
        return 0

    end = dt.date.today()
    try:
        import yfinance as yf
        hist = yf.Ticker(ticker).history(start=since.isoformat(),
                                         end=(end + dt.timedelta(days=1)).isoformat(),
                                         interval="1d", auto_adjust=True)
        series = hist["Close"].dropna() if not hist.empty else None
    except Exception:
        return 0
    if series is None or series.empty:
        return 0

    inserted = 0
    with _connect() as con:
        for idx, price in series.items():
            d = idx.date() if hasattr(idx, "date") else idx
            cur = con.execute(
                "INSERT OR IGNORE INTO prices(trade_date, asset, price) VALUES (?,?,?)",
                (d.isoformat(), asset, float(price)),
            )
            if cur.rowcount:
                inserted += 1
        con.commit()
    _invalidate_caches()
    return inserted


def add_cash_movement(when: dt.date, kind: str, amount: float) -> dict:
    """Record a pure cash movement: DEPOSIT (money into the PEA) or WITHDRAW
    (money out). No ETF, no price/shares. Blocks withdrawing more than held."""
    ensure_seeded()
    kind = kind.upper()
    if kind not in ("DEPOSIT", "WITHDRAW"):
        raise ValueError("kind must be DEPOSIT or WITHDRAW")
    if amount <= 0:
        raise ValueError("amount must be > 0")

    if kind == "WITHDRAW":
        avail = cash_balance_as_of(when)
        if amount > avail + 1e-9:
            raise ValueError(
                f"Retrait refusé : cash disponible au {when.isoformat()} = "
                f"{avail:,.2f} €, impossible de retirer {amount:,.2f} €."
            )

    pf_ccy = current_portfolio_currency()
    with _connect() as con:
        con.execute(
            "INSERT INTO transactions(trade_date,asset,isin,txn_type,price,shares,amount_eur,currency)"
            " VALUES (?,?,?,?,?,?,?,?)",
            (when.isoformat(), "Cash", None, kind, None, None, round(amount, 2), pf_ccy),
        )
        con.commit()
    _invalidate_caches()
    return {"date": when.isoformat(), "type": kind, "amount": round(amount, 2),
            "cash_after": cash_balance_as_of(dt.date.today())}


# ---------------------------------------------------------------------------
# Write APIs — daily prices
# ---------------------------------------------------------------------------

def save_today(prices: dict[str, float],
               shares: dict[str, float] | None = None,
               when: dt.date | None = None) -> dict:
    """Upsert a daily price row into sqlite. Returns a small summary."""
    when = when or dt.date.today()
    ensure_seeded()
    with _connect() as con:
        for asset, price in prices.items():
            if price is None or asset not in ASSETS:
                continue
            con.execute(
                "INSERT OR REPLACE INTO prices(trade_date, asset, price) VALUES (?,?,?)",
                (when.isoformat(), asset, float(price)),
            )
        con.execute("INSERT OR REPLACE INTO meta(key,value) VALUES(?,?)",
                    ("last_save_at", dt.datetime.now().isoformat(timespec="seconds")))
        con.commit()
    sh = shares or shares_held_as_of(when)
    nav = sum(prices.get(a, 0) * sh.get(a, 0) for a in ASSETS)
    return {"date": when.isoformat(), "total_value": round(nav, 2), "rows": len(prices)}


def last_save_at() -> str | None:
    ensure_seeded()
    with _connect() as con:
        row = con.execute("SELECT value FROM meta WHERE key='last_save_at'").fetchone()
    return row["value"] if row else None


# ---------------------------------------------------------------------------
# V11 — FX engine (historical Yahoo rates, cached on disk + in memory)
# ---------------------------------------------------------------------------

# Currency display symbols (used by the UI's money() formatter)
CURRENCY_SYMBOL = {
    "EUR": "€", "USD": "$", "GBP": "£", "JPY": "¥",
    "CHF": "CHF", "CAD": "C$", "AUD": "A$", "CNY": "¥",
}
COMMON_CURRENCIES = ["EUR", "USD", "GBP", "JPY", "CHF", "CAD", "AUD", "CNY"]

# In-memory FX cache: {(date_iso, pair): rate}
_FX_CACHE: dict[tuple[str, str], float] = {}


def _fx_load_cache_from_db() -> None:
    """One-shot load of the fx_rates table into memory (called lazily)."""
    if _FX_CACHE:
        return
    ensure_seeded()
    with _connect() as con:
        rows = con.execute("SELECT trade_date, pair, rate FROM fx_rates").fetchall()
    for r in rows:
        _FX_CACHE[(r["trade_date"][:10], r["pair"])] = float(r["rate"])


def _fx_persist(pair: str, rows: list[tuple[dt.date, float]]) -> None:
    if not rows:
        return
    with _connect() as con:
        for d, rate in rows:
            con.execute(
                "INSERT OR REPLACE INTO fx_rates(trade_date, pair, rate) VALUES (?,?,?)",
                (d.isoformat(), pair, float(rate)),
            )
        con.commit()
    for d, rate in rows:
        _FX_CACHE[(d.isoformat(), pair)] = float(rate)


def _fx_fetch_yahoo(from_ccy: str, to_ccy: str,
                    start: dt.date, end: dt.date) -> list[tuple[dt.date, float]]:
    """Fetch daily FX history from Yahoo. Empty list on failure."""
    if from_ccy == to_ccy:
        return [(start, 1.0)]
    try:
        import yfinance as yf
        # Yahoo uses "{FROM}{TO}=X". E.g. USDEUR=X gives USD/EUR rate
        # (1 USD = X EUR). For ccy pairs with no direct ticker, fall back to
        # the inverse pair and reciprocate.
        symbol = f"{from_ccy}{to_ccy}=X"
        hist = yf.Ticker(symbol).history(
            start=start.isoformat(),
            end=(end + dt.timedelta(days=1)).isoformat(),
            interval="1d", auto_adjust=True,
        )
        if hist.empty:
            # Try the inverse and reciprocate
            inv_symbol = f"{to_ccy}{from_ccy}=X"
            hist = yf.Ticker(inv_symbol).history(
                start=start.isoformat(),
                end=(end + dt.timedelta(days=1)).isoformat(),
                interval="1d", auto_adjust=True,
            )
            if hist.empty:
                return []
            series = (1.0 / hist["Close"].dropna())
        else:
            series = hist["Close"].dropna()
        return [(idx.date(), float(v)) for idx, v in series.items()]
    except Exception:
        return []


def fx_rate(from_ccy: str, to_ccy: str, when: dt.date | None = None) -> float:
    """Return the FX multiplier: 1 unit of `from_ccy` = X units of `to_ccy` on
    `when` (or today). Falls back to the last known rate if the exact date is
    missing (weekends, holidays). Returns 1.0 if from == to. On fetch failure,
    returns 1.0 as a safe fallback (logged via the caller's success message)."""
    from_ccy = (from_ccy or "EUR").upper()
    to_ccy = (to_ccy or "EUR").upper()
    if from_ccy == to_ccy:
        return 1.0
    when = when or dt.date.today()
    pair = f"{from_ccy}{to_ccy}"
    _fx_load_cache_from_db()

    # Exact date hit
    hit = _FX_CACHE.get((when.isoformat(), pair))
    if hit is not None:
        return hit

    # Most recent earlier date (forward-fill)
    candidates = [k for k in _FX_CACHE if k[1] == pair and k[0] <= when.isoformat()]
    if candidates:
        return _FX_CACHE[max(candidates)]

    # Cold cache: fetch a window ending today, persist, then look up
    start = min(when, dt.date.today() - dt.timedelta(days=400))
    rows = _fx_fetch_yahoo(from_ccy, to_ccy, start, dt.date.today())
    if rows:
        _fx_persist(pair, rows)
        # Try again
        candidates = [k for k in _FX_CACHE if k[1] == pair and k[0] <= when.isoformat()]
        if candidates:
            return _FX_CACHE[max(candidates)]
        # If `when` is older than the earliest fetched, use the earliest
        all_dates = sorted(k[0] for k in _FX_CACHE if k[1] == pair)
        if all_dates:
            return _FX_CACHE[(all_dates[0], pair)]
    # Last-ditch safe fallback
    return 1.0


def prefetch_fx_window(from_ccy: str, to_ccy: str,
                       start: dt.date, end: dt.date) -> None:
    """Warm the FX cache for a date window. Cheap no-op if already covered.
    Used before nav_series so per-date lookups are guaranteed O(1)."""
    if from_ccy == to_ccy:
        return
    pair = f"{from_ccy.upper()}{to_ccy.upper()}"
    _fx_load_cache_from_db()
    earliest = min((dt.date.fromisoformat(k[0]) for k in _FX_CACHE if k[1] == pair),
                   default=None)
    latest = max((dt.date.fromisoformat(k[0]) for k in _FX_CACHE if k[1] == pair),
                 default=None)
    need_fetch_start = start if (earliest is None or earliest > start) else None
    need_fetch_end = end if (latest is None or latest < end) else None
    if need_fetch_start or need_fetch_end:
        rows = _fx_fetch_yahoo(from_ccy.upper(), to_ccy.upper(),
                               need_fetch_start or start,
                               need_fetch_end or end)
        if rows:
            _fx_persist(pair, rows)


def latest_price_date() -> dt.date | None:
    ensure_seeded()
    with _connect() as con:
        row = con.execute("SELECT MAX(trade_date) AS d FROM prices").fetchone()
    if not row or not row["d"]:
        return None
    return dt.date.fromisoformat(row["d"][:10])


def existing_price_dates(start: dt.date | None = None) -> set[dt.date]:
    ensure_seeded()
    sql = "SELECT DISTINCT trade_date FROM prices"
    params: tuple = ()
    if start is not None:
        sql += " WHERE trade_date >= ?"
        params = (start.isoformat(),)
    with _connect() as con:
        rows = con.execute(sql, params).fetchall()
    return {dt.date.fromisoformat(r["trade_date"][:10]) for r in rows}


def backfill_prices(rows: list[tuple[dt.date, str, float]]) -> dict:
    if not rows:
        return {"inserted": 0, "skipped": 0, "dates": []}
    ensure_seeded()
    inserted = 0
    inserted_dates: set[dt.date] = set()
    with _connect() as con:
        for d, asset, price in rows:
            if asset not in ASSETS or price is None:
                continue
            cur = con.execute(
                "INSERT OR IGNORE INTO prices(trade_date, asset, price) VALUES (?,?,?)",
                (d.isoformat(), asset, float(price)),
            )
            if cur.rowcount:
                inserted += 1
                inserted_dates.add(d)
        con.commit()
    return {"inserted": inserted, "skipped": len(rows) - inserted,
            "dates": sorted(inserted_dates)}


def price_on_date(asset: str, when: dt.date) -> tuple[float | None, str]:
    """VA2: return (price_native, source) for `asset` at `when`.

    Lookup order:
      1. local `prices` table (cached close, native currency)
      2. Yahoo `fetch_history(when, when)` (single bar)
      3. nearest earlier trading day in local DB (forward-fill fallback)
      4. (None, "not_found") if absolutely nothing available

    `source` is one of: "db_cache" | "yahoo_fetch" | "fallback_prev_close" | "not_found".
    The price is always in the asset's NATIVE currency (FX conversion is
    the caller's responsibility).
    """
    if asset not in ASSETS:
        return None, "not_found"
    ensure_seeded()
    with _connect() as con:
        row = con.execute(
            "SELECT price FROM prices WHERE asset=? AND trade_date=?",
            (asset, when.isoformat()),
        ).fetchone()
    if row is not None:
        return float(row["price"]), "db_cache"

    # Try Yahoo single-day fetch
    try:
        import prices as _prices
        hist = _prices.fetch_history(when, when)
        rows = hist.get(asset) or []
        if rows:
            d, p = rows[0]
            if p is not None:
                # Cache it
                with _connect() as con:
                    con.execute(
                        "INSERT OR REPLACE INTO prices(trade_date, asset, price) VALUES(?,?,?)",
                        (d.isoformat(), asset, float(p)),
                    )
                    con.commit()
                _invalidate_caches()
                return float(p), "yahoo_fetch"
    except Exception:
        pass

    # Forward-fill from the nearest earlier saved trading day
    with _connect() as con:
        row = con.execute(
            "SELECT price FROM prices WHERE asset=? AND trade_date<=?"
            " ORDER BY trade_date DESC LIMIT 1",
            (asset, when.isoformat()),
        ).fetchone()
    if row is not None:
        return float(row["price"]), "fallback_prev_close"
    return None, "not_found"


def refetch_recent_closes(days: int = 7) -> dict:
    """V13.1: re-fetch the last `days` calendar days via fetch_history and
    OVERWRITE existing rows. This kills any intraday snapshot that the
    "Refresh & Save" button (or a crashed cron) may have written during
    market hours — Yahoo's `history` endpoint always returns the final
    close for past trading days. Idempotent. Returns a small summary.

    Why this exists: `save_today` writes whatever Yahoo serves *right now*,
    which is the live intraday quote when called during market hours. The
    18:00 cron used to skip overwriting if the date was already present,
    leaving stale intraday values forever. This fix-up pass closes that hole.
    """
    import prices as _prices  # local import to avoid circular dep at module load
    today = dt.date.today()
    start = today - dt.timedelta(days=days)
    try:
        history = _prices.fetch_history(start, today)
    except Exception:
        return {"overwritten": 0, "dates": [], "error": "yahoo_fetch_failed"}
    ensure_seeded()
    overwritten = 0
    touched: set[dt.date] = set()
    with _connect() as con:
        for asset, series in history.items():
            if asset not in ASSETS:
                continue
            for d, p in series:
                if p is None:
                    continue
                con.execute(
                    "INSERT OR REPLACE INTO prices(trade_date, asset, price) VALUES (?,?,?)",
                    (d.isoformat(), asset, float(p)),
                )
                overwritten += 1
                touched.add(d)
        con.commit()
    return {"overwritten": overwritten, "dates": sorted(touched)}


# ---------------------------------------------------------------------------
# Unitized (time-weighted) VL engine
# ---------------------------------------------------------------------------

def compute_vl_series(price_history: pd.DataFrame | None = None,
                      inception: dt.date | None = None) -> pd.DataFrame:
    """Day-by-day NAV, fund units, VL (base 100 at inception) and net invested.

    Units are created when you BUY (cash in) and destroyed when you SELL (cash
    out), at the VL prevailing just before the flow — so VL reflects ONLY market
    performance, not deposits/withdrawals. With no post-inception flows this
    reduces exactly to NAV / NAV_inception * 100.
    Returns columns: date, nav, units, vl, net_invested.
    """
    if price_history is None:
        price_history = load_price_history()
    inception = inception or get_inception_date()
    if price_history.empty:
        return pd.DataFrame(columns=["date", "nav", "units", "vl", "net_invested"])

    nv = nav_series(price_history, inception)
    if nv.empty:
        return pd.DataFrame(columns=["date", "nav", "units", "vl", "net_invested"])

    dates = [d.date() for d in nv["date"]]
    assets = [a for a in ASSETS if a in price_history.columns]
    ph = price_history[price_history["date"].dt.date >= inception].sort_values("date").reset_index(drop=True)

    walk = _cash_walk()
    deposit_by_date = walk["deposit_by_date"]
    # External deposits AFTER inception drive unit creation. Inception-day
    # deposits are the initial capital (folded into units_0). Map each deposit
    # to the first trading date >= its date.
    cf_on: dict[dt.date, float] = defaultdict(float)
    for d_dep, amt in deposit_by_date.items():
        if d_dep <= inception:
            continue
        td = next((x for x in dates if x >= d_dep), dates[-1] if dates else None)
        if td is not None:
            cf_on[td] += amt

    def etf_value_on(row_idx: int, held: dict[str, float]) -> float:
        total = 0.0
        for a in assets:
            v = ph.at[row_idx, a]
            if pd.notna(v):
                total += float(v) * held.get(a, 0)
        return total

    out_rows = []
    units = None
    for i, d in enumerate(dates):
        nav = float(nv.at[i, "nav"])
        if i == 0:
            units = (nav / 100.0) if nav > 0 else 1.0
        else:
            cf = cf_on.get(d, 0.0)
            if cf and units:
                # VL just before the deposit: today's prices on yesterday's
                # shares + yesterday's cash (NAV excluding the new money).
                held_prev = shares_held_as_of(dates[i - 1])
                cash_prev = float(nv.at[i - 1, "cash"])
                nav_pre = etf_value_on(i, held_prev) + cash_prev
                vl_pre = (nav_pre / units) if units else 100.0
                if vl_pre > 0:
                    units += cf / vl_pre
        vl = (nav / units) if units else 100.0
        net_inv = external_deposits_as_of(d)
        out_rows.append({"date": nv.at[i, "date"], "nav": nav, "units": units,
                         "vl": vl, "net_invested": net_inv})

    return pd.DataFrame(out_rows)


# ---------------------------------------------------------------------------
# Derived KPI snapshot
# ---------------------------------------------------------------------------

def compute_snapshot(static: PortfolioStatic,
                     prices: dict[str, float],
                     price_history: pd.DataFrame) -> dict:
    """Build the KPI snapshot consumed by the UI.

    Uses TIME-VARYING shares and a unitized VL. `prices` is the current
    per-asset price (live or last close); daily P&L is measured on the shares
    held going into the latest day, so a same-day trade doesn't masquerade as
    performance.
    """
    inception = get_inception_date()
    shares_now = static.shares  # current (today) shares
    cash_now = cash_balance_as_of(dt.date.today())

    # V11: FX conversion to the active portfolio's base currency
    pf_ccy = current_portfolio_currency()
    today = dt.date.today()

    def _to_pf(amount: float, native_ccy: str, when: dt.date | None = None) -> float:
        native_ccy = (native_ccy or "EUR").upper()
        if native_ccy == pf_ccy:
            return amount
        return amount * fx_rate(native_ccy, pf_ccy, when or today)

    positions: dict[str, dict] = {}
    etf_value = 0.0
    for asset in ASSETS:
        price = prices.get(asset)
        sh = shares_now.get(asset, 0)
        if price is None:
            continue
        native_ccy = (static.currencies.get(asset) or "EUR").upper()
        rate = 1.0 if native_ccy == pf_ccy else fx_rate(native_ccy, pf_ccy, today)
        native_value = price * sh
        value = native_value * rate
        positions[asset] = {
            "price": price, "shares": sh, "value": value,
            "native_price": price, "native_value": native_value,
            "currency": native_ccy, "fx_rate": rate,
        }
        etf_value += value
    total_value = etf_value + cash_now

    # Cash (PEA) appears as its own allocation line when non-zero. It has no
    # price/shares/returns; it's pure liquidity sitting in the plan.
    if cash_now > 1e-9:
        positions["Cash"] = {
            "price": None, "shares": 0, "value": cash_now, "is_cash": True,
            "daily_return": 0.0, "daily_pnl": 0.0,
            "total_return": 0.0, "total_return_eur": 0.0,
            "inception_price": None, "target_allocation": 0.0, "drift": 0.0,
        }

    # "Yesterday" = second-to-last price row; daily P&L on shares held into it.
    prev_row = None
    prev_date = None
    if not price_history.empty:
        hist_sorted = price_history.sort_values("date").reset_index(drop=True)
        if len(hist_sorted) >= 2:
            prev_row = hist_sorted.iloc[-2]
            prev_date = prev_row["date"].date()
    shares_prev = shares_held_as_of(prev_date) if prev_date else shares_now

    daily_pnl_eur = 0.0
    prev_nav = 0.0
    for asset in ASSETS:
        cur = prices.get(asset)
        prev = None
        if prev_row is not None and asset in prev_row.index and pd.notna(prev_row[asset]):
            prev = float(prev_row[asset])
        sh_prev = shares_prev.get(asset, 0)
        p = positions.setdefault(asset, {"price": None, "shares": shares_now.get(asset, 0), "value": 0.0})
        native_ccy = (static.currencies.get(asset) or "EUR").upper()
        if cur is not None and prev:
            # daily_return = NATIVE price ratio (currency cancels) — measures
            # pure stock performance independent of FX moves.
            p["daily_return"] = (cur / prev) - 1.0
            # FX rates: today vs previous trade date
            r_today = 1.0 if native_ccy == pf_ccy else fx_rate(native_ccy, pf_ccy, today)
            r_prev = (1.0 if native_ccy == pf_ccy
                      else fx_rate(native_ccy, pf_ccy, prev_date or today))
            # daily_return_pf = portfolio-currency ratio (stock + FX combined).
            # Equal to daily_return when native_ccy == pf_ccy.
            p["daily_return_pf"] = ((cur * r_today) / (prev * r_prev)) - 1.0
            # Daily P&L IN PORTFOLIO CURRENCY: today's price × FX_today minus
            # yesterday's price × FX_yesterday, on shares held into today.
            p["daily_pnl"] = (cur * r_today - prev * r_prev) * sh_prev
            daily_pnl_eur += p["daily_pnl"]
            prev_nav += prev * r_prev * sh_prev
        else:
            p["daily_return"] = 0.0
            p["daily_return_pf"] = 0.0
            p["daily_pnl"] = 0.0

    daily_pnl_pct = (daily_pnl_eur / prev_nav) if prev_nav else 0.0
    for p in positions.values():
        p["allocation"] = (p["value"] / total_value) if total_value else 0.0

    # --- Unitized VL series (time-weighted performance) ---
    vl_series = compute_vl_series(price_history, inception)
    if not vl_series.empty:
        last = vl_series.iloc[-1]
        vl = float(last["vl"])
        net_invested = float(last["net_invested"])
        # If live `prices` differ from the last stored row, re-value NAV with
        # live prices but keep the units from the series.
        units = float(last["units"]) if last["units"] else None
        if units and total_value > 0:
            vl = total_value / units
        nav_for_vl = total_value if total_value > 0 else float(last["nav"])
    else:
        vl = 100.0
        net_invested = sum((t["amount"] if t["type"] == "BUY" else -t["amount"])
                           for t in _all_transactions())
        nav_for_vl = total_value

    total_return_pct = vl / 100.0 - 1.0
    # Cash P&L = current market value − net cash invested (realized + unrealized)
    cash_pnl_eur = total_value - net_invested

    # --- Per-asset total return + target (inception) allocation ---
    inception_nav = 0.0
    if not price_history.empty:
        first = price_history.sort_values("date").iloc[0]
        first_date = first["date"].date()
        shares_first = shares_held_as_of(first_date)
        # Inception NAV in PORTFOLIO currency
        inception_nav = 0.0
        for a in ASSETS:
            if a not in first.index or not pd.notna(first[a]):
                continue
            a_ccy = (static.currencies.get(a) or "EUR").upper()
            r_inc = 1.0 if a_ccy == pf_ccy else fx_rate(a_ccy, pf_ccy, first_date)
            inception_nav += float(first[a]) * shares_first.get(a, 0) * r_inc

        for asset in ASSETS:
            cur = prices.get(asset)
            inc_price = float(first[asset]) if asset in first.index and pd.notna(first[asset]) else None
            sh = shares_now.get(asset, 0)
            a_ccy = (static.currencies.get(asset) or "EUR").upper()
            r_today = 1.0 if a_ccy == pf_ccy else fx_rate(a_ccy, pf_ccy, today)
            r_inc = 1.0 if a_ccy == pf_ccy else fx_rate(a_ccy, pf_ccy, first_date)
            p = positions.setdefault(asset, {"price": None, "shares": sh, "value": 0.0})
            if cur is not None and inc_price:
                # Total return = portfolio-currency price change (incl. FX move)
                p["total_return"] = (cur * r_today) / (inc_price * r_inc) - 1.0
                p["total_return_eur"] = (cur * r_today - inc_price * r_inc) * sh
                p["inception_price"] = inc_price
            else:
                p["total_return"] = 0.0
                p["total_return_eur"] = 0.0
                p["inception_price"] = None
            if inc_price is not None and inception_nav > 0:
                p["target_allocation"] = (inc_price * r_inc * shares_first.get(asset, 0)) / inception_nav
            else:
                p["target_allocation"] = 0.0
            p["drift"] = p.get("allocation", 0.0) - p["target_allocation"]

    # Drop fully-exited assets (0 shares and 0 value) from the snapshot view
    positions = {a: p for a, p in positions.items()
                 if not (abs(p.get("shares", 0)) < 1e-9 and abs(p.get("value", 0)) < 1e-9)}

    return {
        "positions": positions,
        "total_value": total_value,
        "daily_pnl_eur": daily_pnl_eur,
        "daily_pnl_pct": daily_pnl_pct,
        "prev_total": prev_nav,
        "inception_nav": inception_nav,
        "inception_date": inception,
        "total_return_eur": cash_pnl_eur,     # now cash-based (NAV − net invested)
        "total_return_pct": total_return_pct, # time-weighted (VL-based)
        "net_invested": net_invested,
        "cash_pnl_eur": cash_pnl_eur,
        "cash_balance": cash_now,
        "etf_value": etf_value,
        "vl": vl,
    }
